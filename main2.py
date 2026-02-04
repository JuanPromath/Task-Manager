
from openpyxl import load_workbook
import util as u

arquivo = load_workbook("tudo modelado.xlsx")
arquivoR = load_workbook("tudo modelado.xlsx",data_only=True, read_only=True,)

print(arquivo.sheetnames)

def criar_ciclo():
    ciclo = {}
    r = input("Digite o nome do ciclo: ")
    ciclo['nome'] = r
    r = input("Desejar digitar o tempo em decimal ou hh:mm:ss :[d][h]\nR: ")
    if r == 'd':
        hora = float(input("Digite a hora a ser mostrada: "))
        ciclo['tempo total'] = u.objetoHorasParaStringHoras(u.horaParaHorasBonita(hora))
        ciclo['tempo total decimal'] = hora
    elif r == 'h':
        hora = u.criarHora()
        ciclo['tempo total'] = u.objetoHorasParaStringHoras(hora)
        ciclo['tempo total decimal'] = u.lessDecimalBadHour(u.converteParaDecimalFeio(hora),hora)
    else:
        print('digite uma entrada valida')
    print(ciclo)
    insert(ciclo, 'Ciclo')

def criar_atividade():
    atividade = {}
    r = input('Digite o nome da atividade: ')
    atividade['Nome'] = r
    insert(atividade, 'Atividade')

def table2List(tab):
    Tab = arquivoR[tab]
    header = Tab['1']
    qtdLines = len(Tab['A'])
    lista = []
    for i in range(1, qtdLines):
        res = {}
        for j, columnHeader in enumerate(header):
            valorCelula = Tab[f'{chr(ord('A') + j)}{i + 1}'].value 
            print(Tab[f'{chr(ord('A') + j)}{i + 1}'].internal_value)
            res[columnHeader.value] = Tab[f'{chr(ord('A') + j)}{i + 1}'].value
        lista.append(res)
    
    return lista
        
def insert(inserted, tab):
    archiveTab = arquivo[tab]
    codigos = archiveTab['A']
    line = len(codigos)
    print(f'A{line + 1}')
    StartColumn = ord('B')

    archiveTab[f'A{line + 1}'] = line

    for key in inserted.keys():
        #print(key)
        #print(f'{StartColumn} - {chr(StartColumn)} - {chr(StartColumn)}{line + 1} - {cicloDict[key]}')
        archiveTab[f'{chr(StartColumn)}{line + 1}'] = inserted[key]
        StartColumn += 1
    for cell in codigos:
        print(cell.value)
    arquivo.save('tudo modelado.xlsx')

def insert_cycle(cicloDict):
    #achar a ultima linha
    #gerar um código unico
    cicloTab = arquivo['Ciclo']
    codigos = cicloTab['A']
    line = len(codigos)
    print(f'A{line + 1}')
    StartColumn = ord('B')

    cicloTab[f'A{line + 1}'] = line

    for key in cicloDict.keys():
        #print(key)
        #print(f'{StartColumn} - {chr(StartColumn)} - {chr(StartColumn)}{line + 1} - {cicloDict[key]}')
        cicloTab[f'{chr(StartColumn)}{line + 1}'] = cicloDict[key]
        StartColumn += 1
    for cell in codigos:
        print(cell.value)
    arquivo.save('tudo modelado.xlsx')

def gerenciar(ciclo):
    print(ciclo)
    gerenciando = True
    while gerenciando:
        showList(table2List('uniao ciclo atividade'))
        gerenciando = False

def gerenciarCiclo():
    #listar os ciclos
    list = table2List('Ciclo')
    for i,line in enumerate(list):
        res = f'[{i}] - '
        for pos, key in enumerate(line):
            if pos == len(line.keys()) - 1:
                res += f'{key}: {line[key]}'
            else:
                res += f'{key}: {line[key]} - '

        print(res)
    #input pra escolher ciclo
    r = input('R: ')
    gerenciar(list[int(r)]['tempo total'])
    #sistema menu de ciclo

def showList(list):
    for line in list:
        res = ''
        for pos, key in enumerate(line):
            if pos == len(line.keys()) - 1:
                res += f'{key}: {line[key]}'
            else:
                res += f'{key}: {line[key]} - '
        print(res)

while True:
    
    r = input("[0] - criar ciclo\n[1] - Criar atividade\n[2] - listar ciclos\n[3] - gerenciar ciclo\n[4] - encerrar aplicação\nR: ")

    if r == "0":
        criar_ciclo()
    elif r == '1':
        criar_atividade()
    elif r == '2':
        showList(table2List('Ciclo'))
    elif r == '3':
        gerenciarCiclo()
    elif r == "4":
        print('fechado')
        arquivo.close()
        break
